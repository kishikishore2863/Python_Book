import openpyxl

wb = openpyxl.load_workbook('Master.xlsx')
ws = wb.active
ws = wb['Lectures']
print(ws)

print(ws.max_row)
print(ws.max_column)

# values = [[ws.cell(row=j,column=i).value for i in range(1,ws.max_column+1)]for j in range(1,ws.max_row+1)]
# for i in values:
#     print(i)
listvalue = list()
for value in ws.iter_rows(
    min_row =1,max_row=11,min_col=1,max_col=5,values_only=True
):
    listvalue.append(value)
print(listvalue)