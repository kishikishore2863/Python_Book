from openpyxl import load_workbook
from openpyxl.workbook import Workbook

wb = load_workbook('emp_data.xlsx')
ws = wb.active



dic ={}
for row in range(2,ws.max_row+1):
    basic = ws[f'c{row}'].value
    empId = ws[f'a{row}'].value
    dic[empId]=basic
print(dic)

newGross={}
for k,v in dic.items():
    newGross[k]=(v*0.1)+(v*0.18)

newBook = Workbook()
newSheet =newBook.active
newSheet['a1'] = "empId"
newSheet['b1'] = "Name"
newSheet['c1'] = "GrossSalary"
index=2
for k,v in newGross.items():
    newSheet[f"a{index}"] = k
    newSheet[f"b{index}"] = ws[f'b{index}'].value
    newSheet[f"c{index}"] = v
    index =index+1
newBook.save('emp_salary.xlsx')
