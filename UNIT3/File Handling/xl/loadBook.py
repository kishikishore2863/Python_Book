from openpyxl import load_workbook,Workbook
# wb = Workbook()
# ws = wb.active
#
# ws["A1"] = "Name"
# ws["B1"] = "Marks"
#
# wb.save("data.xlsx")
# print("✅ Fresh valid Excel file created")
#
wb = load_workbook('data.xlsx')
ws = wb.active

ws["A2"] = "Kishore"
ws["A3"] = "jatin"
ws["B2"] = 95
ws["B3"] = 100
ws["C1"] = "Loop"
for i in range(2,101):
    ws[f"C{i}"] = "Kishore"

wb.save("data.xlsx")
