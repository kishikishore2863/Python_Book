from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = 'Student'

ws['A1']= 'Name'
ws['B1'] = 'Marks'

wb.save('data.xlsx')



