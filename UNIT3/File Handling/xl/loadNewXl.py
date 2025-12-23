from openpyxl import load_workbook

wb = load_workbook('newXlData.xlsx')
ws = wb.active

ws['A1'] = 'Subjects'

ws['C1'] = 'Max Marks'

ws['A2'] = 'Python'
ws['A3'] = 'OOSE'
ws['A4'] = 'OSD'
ws['A5'] = 'DBA'
ws['A6'] = 'Data Structures'

for i in range(2,7):
    ws[f'C{i}'] =40

ws['B1'] = "Marks Obtained"
ws['B2'] = 25
ws['B3'] = 25
ws['B4'] = 25
ws['B5'] = 20
ws['B6'] = 24



wb.save('newXlData.xlsx')