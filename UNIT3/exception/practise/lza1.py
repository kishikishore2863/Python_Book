import datetime

import openpyxl


wb = openpyxl.load_workbook('Master.xlsx')
print(type(wb))

ws  = wb.active
print(type(ws))


ws = wb['Lectures']
headers = [ ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(headers)
rows = []
for i in range(1,ws.max_row):
    row = []
    for j in range(1,ws.max_column):
        row.append(ws.cell(row=i,column=j).value)
    rows.append(row)

